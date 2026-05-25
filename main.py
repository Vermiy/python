from faster_whisper import WhisperModel
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openai import OpenAI
import os
import json
from datetime import datetime


def load_env_file(env_path: str | None = None):
    path = Path(env_path) if env_path else Path(__file__).resolve().parent / ".env"
    if not path.exists():
        return

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        os.environ.setdefault(key, value)


load_env_file()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY");
client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None;

EXCEL_FILE = "calls.xlsx"


model = WhisperModel("small", device="cpu")
audio_path = Path("audio/2025-05-28_14-46_0975668742_incoming (2).mp3")

ESTIMATE_MANAGER_AI_PROMPT = """
Ти — AI асистент для контролю якості роботи менеджерів автосервісу.

Тобі буде передано транскрипт телефонної розмови менеджера з клієнтом.

Твоє завдання:
1. Проаналізувати діалог.
2. Визначити якість роботи менеджера.
3. Заповнити поля оцінки.
4. Бути максимально об'єктивним.
5. Якщо інформації немає — ставити "Невідомо".

ВАЖЛИВО:
- Оцінюй лише те, що реально є у діалозі.
- Не вигадуй інформацію.
- Якщо менеджер поводиться грубо, неуважно, перебиває клієнта,
  не намагається допомогти або порушує стандарти спілкування —
  ОБОВ'ЯЗКОВО вкажи це у полі "comment".
- Якщо є критичні проблеми у розмові —
  починай comment з "❌ ПРОБЛЕМНИЙ ДЗВІНОК:"
- Якщо дзвінок хороший —
  можеш починати comment з "✅"

Правила оцінки:
- 1 = виконано
- 0 = не виконано
- "Невідомо" = неможливо визначити

Потрібно повернути JSON строго у такому форматі:

{
  "call_type": "",
  "phone_number": "",
  "branch": "",
  "manager": "",

  "greeting": 0,
  "asked_car_body": 0,
  "asked_car_year": 0,
  "asked_mileage": 0,
  "offered_full_diagnostics": 0,
  "asked_previous_repairs": 0,
  "service_booking_date": "",
  "goodbye": 0,

  "top_100_work": "",
  "followed_top_100_instructions": "Так",
  "missed_top_100_recommendations": "",

  "result": "",
  "score": 0,
  "parts": "",

  "comment": ""
}

Логіка score:
Порахуй кількість пунктів де менеджер отримав 1:
- greeting
- asked_car_body
- asked_car_year
- asked_mileage
- offered_full_diagnostics
- asked_previous_repairs
- goodbye

score = сума всіх одиниць.

Правила для comment:
- Коротко і по суті.
- Якщо були проблеми — детально вкажи які саме.
- Вказуй:
  - грубість
  - невпевненість
  - відсутність уточнюючих питань
  - відсутність запису
  - погану комунікацію
  - ігнорування потреб клієнта
  - порушення скрипту
- Якщо менеджер добре відпрацював — теж коротко зазнач.

Транскрипт буде переданий нижче.
"""


HEADERS = [
    "Дата",
    "Тип звернення",
    "Номер телефону",
    "Філія",
    "Менеджер",
    "Початок розмови, представлення",
    "Чи дізнвся менеджер кузов атвомобіля",
    "Чи дізнався менеджер рік автомобіля",
    "Чи дізнався менеджр пробіг",
    "Пропозиція про комплексну діагностику",
    "Дізнався які роботи робилися раніше",
    "Запис на сервіс, Дата",
    "Завершення розмови прощання",
    "Яка робота з топ 100",
    "Чи дотримувався всіх інструкцій з топ 100 робіт Да/Ні",
    "Яких рекоменадцій менеджер не дотримувався з топ 100 робіт",
    "Результат",
    "Оцінка",
    "Запчастини",
    "Коментар",
]

def extract_text_from_call(path: str):
    segments, info = model.transcribe(
    path,
    beam_size=5,
    language="uk"
)
    return segments, info;

def estimate_manager_response(call_context: str):
    if client is None:
        print("OPENAI_API_KEY is not set. Skipping AI evaluation.")
        return None

    ctx = ESTIMATE_MANAGER_AI_PROMPT + call_context;
    response = client.responses.create(
    model="gpt-5-mini",
    input=ctx
)
    print(response.output_text);
    return response.output_text


def init_excel():
    path = Path(EXCEL_FILE)

    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        for col, header in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=col, value=header)

        wb.save(EXCEL_FILE)

        print("Excel file created with headers")
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    existing_headers = [
        ws.cell(row=1, column=col).value
        for col in range(1, ws.max_column + 1)
    ]

    changed = False

    for header in HEADERS:
        if header not in existing_headers:
            ws.cell(
                row=1,
                column=ws.max_column + 1,
                value=header
            )
            changed = True
            print(f"Added missing header: {header}")

    if changed:
        wb.save(EXCEL_FILE)
        print("Excel updated")
    else:
        print("All headers already exist")

def parse_ai_json_response(response_text: str):
    if not response_text:
        return None

    try:
        return json.loads(response_text)
    except json.JSONDecodeError:
        start = response_text.find("{")
        end = response_text.rfind("}")
        if start == -1 or end == -1 or end <= start:
            return None

        candidate = response_text[start:end + 1]
        try:
            return json.loads(candidate)
        except json.JSONDecodeError:
            return None


def get_phone_from_audio_filename(file_path: Path):
    stem = file_path.stem
    if "_incoming" not in stem:
        return None

    before_incoming = stem.split("_incoming", 1)[0]
    if not before_incoming:
        return None

    return before_incoming.rsplit("_", 1)[-1] or None


def write_call_result_in_xl(ai_response_text: str, call_audio_path: Path):
    data = parse_ai_json_response(ai_response_text)
    if data is None:
        print("AI response is not valid JSON. Skipping Excel write.")
        return

    phone_from_filename = get_phone_from_audio_filename(call_audio_path)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        data.get("call_type", "Невідомо"),
        phone_from_filename or data.get("phone_number", "Невідомо"),
        data.get("branch", "Невідомо"),
        data.get("manager", "Невідомо"),
        data.get("greeting", "Невідомо"),
        data.get("asked_car_body", "Невідомо"),
        data.get("asked_car_year", "Невідомо"),
        data.get("asked_mileage", "Невідомо"),
        data.get("offered_full_diagnostics", "Невідомо"),
        data.get("asked_previous_repairs", "Невідомо"),
        data.get("service_booking_date", "Невідомо"),
        data.get("goodbye", "Невідомо"),
        data.get("top_100_work", "Невідомо"),
        data.get("followed_top_100_instructions", "Невідомо"),
        data.get("missed_top_100_recommendations", "Невідомо"),
        data.get("result", "Невідомо"),
        data.get("score", "Невідомо"),
        data.get("parts", "Невідомо"),
        data.get("comment", "Невідомо"),
    ]

    ws.append(row)
    wb.save(EXCEL_FILE)
    print("Call analysis written to Excel")


segments, info = extract_text_from_call(str(audio_path));


full_text = ""
for segment in segments:
    text = segment.text.strip()
    full_text += text + "\n"


init_excel();
ai_response_text = estimate_manager_response(full_text);
if ai_response_text:
    write_call_result_in_xl(ai_response_text, audio_path)
